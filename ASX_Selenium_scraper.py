from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service

import openpyxl
import time

# List of ASX codes for which EOD sellement prices will be obtained. Adjust as desired.
ASXshareslist = [['AC8',0],['AGL',0],['ALK',0],['ANZ',0],['BHP',0],
                 ['CAN',0],['CBA',0],['VUK',0],['DHG',0],['ERA',0],
                 ['NEC',0],['IAG',0],['JHG',0],['MPL',0],['MQG',0],
                 ['MYR',0],['NAB',0],['NHF',0],['PAN',0],['PTR',0],
                 ['RIO',0],['RNE',0],['S32',0],['SCP',0],['WBC',0],
                 ['WOW',0],['WTC',0],['XRG',0],['AMP',0],['AKE',0],
                 ['ASM',0],['VML',0]]


#,['ALK',0]

# location of Chrome driver. Adjust as necessary.
#driver = webdriver.Chrome('C:\\Users\\roman\\OneDrive\\GitHub\\ASXscraper\\chromedriver.exe')
s = Service('C:\\Users\\roman\\OneDrive\\GitHub\\ASXscraper\\chromedriver.exe')
driver = webdriver.Chrome(service=s)

# location of Excel workbook where price data is saved. Adjust as necessary.
fn = "C:\\Users\\roman\\OneDrive\\GitHub\\ASXscraper\\Shares.xlsx"

# open above workbook so that data can be added
wb = openpyxl.load_workbook(fn)

# Worksheet assumed to exist in Above Excel workbook
ws = wb['Shares'] 

i=2
for x in ASXshareslist:
    i = i+1

    # ASX share code being considered
    sShare = x[0]

    # loops through ASX web pages that contain EOD share prices
    driver.get('https://www2.asx.com.au/markets/company/'+sShare)
    # need delay to alLow time for web page to fully load
    time.sleep(5)
    # this is the 'magik' tag that finds the string that contains the price (and some other data)
    # el = driver.find_element_by_tag_name("dd").text
    
    el = driver.find_element(By.TAG_NAME, "dd").text
    ela = el.split()
    
    # ASX EOD share price is collected and formatted
    # price is the first piece of 'blank' seperated data
    fPrice = float(ela[0])
    
    # price saved here for potential future code enhancements
    x[1] = fPrice
    sPrice = '{:7.3f}'.format(fPrice)
    print(sShare+' : ',sPrice)
    #print(sShare+' : ',el)

    # populating Excel workbook cells for considered share
    # This code can be expanded/modified as desired.
    cr = ws.cell(row=i, column=2)
    cr.value=sShare
    cr = ws.cell(row=i, column=3)
    cr.value=fPrice

# Save Excel workbook
wb.save(fn)

# close interface to Chrome browser
driver.close()
